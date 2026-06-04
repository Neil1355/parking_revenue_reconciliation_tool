"""
Report Module - Excel Report Generation

Generates summary and detailed reports from ledger data and exports to
formatted Excel workbooks.
"""

from typing import Dict, List, Any, Optional
from datetime import datetime
from pathlib import Path

# Optional dependencies used when exporting to Excel
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
except Exception:  # pragma: no cover - optional at runtime
    Workbook = None
    Font = None
    Alignment = None
    PatternFill = None
    get_column_letter = None

# pandas is used in several type annotations and helper functions; import if available
try:
    import pandas as pd
except Exception:  # pragma: no cover - optional at runtime
    pd = None

def generate_summary_report(
    ledger_entries: List[Dict[str, Any]]
) -> Dict[str, Any]:
    """
    Create a summary report with aggregate statistics.

    Args:
        ledger_entries: List of aggregated ledger entries

    Returns:
        Dictionary containing:
        - total_revenue: Sum of all ledger amounts
        - total_transactions: Count of all transactions
        - entry_count: Number of ledger entries
        - period_start: Earliest date in data
        - period_end: Latest date in data
        - average_transaction_value: Mean transaction amount

    Logic Steps:
        1. Initialize summary statistics dictionary
        2. Calculate total revenue across all entries
        3. Calculate total transaction count
        4. Extract date range from entries
        5. Calculate average transaction value
        6. Return summary dictionary
    """
    # Initialize
    total_revenue = 0.0
    total_transactions = 0
    entry_count = len(ledger_entries or [])
    dates = []

    for entry in (ledger_entries or []):
        # Amount extraction: support multiple common field names
        amount = None
        for key in ('total_revenue', 'revenue_amount', 'amount', 'value'):
            if key in entry and entry[key] is not None:
                try:
                    amount = float(entry[key])
                except Exception:
                    amount = None
                break
        if amount is not None:
            total_revenue += amount

        # Transaction count if present on entry
        for cnt_key in ('transaction_count', 'transactions', 'count'):
            if cnt_key in entry and entry[cnt_key] is not None:
                try:
                    total_transactions += int(entry[cnt_key])
                except Exception:
                    pass

        # Date extraction
        for dkey in ('transaction_date', 'date', 'period'):
            if dkey in entry and entry[dkey] is not None:
                dval = entry[dkey]
                try:
                    if isinstance(dval, datetime):
                        dates.append(dval.date())
                    elif isinstance(dval, str):
                        # try ISO formats first
                        try:
                            dates.append(datetime.fromisoformat(dval).date())
                        except Exception:
                            try:
                                dates.append(datetime.strptime(dval, '%Y-%m-%d').date())
                            except Exception:
                                pass
                    else:
                        # numeric timestamp
                        try:
                            dates.append(datetime.fromtimestamp(float(dval)).date())
                        except Exception:
                            pass
                except Exception:
                    pass
                break

    if total_transactions == 0:
        # Fallback: if no per-entry transaction counts, treat each ledger entry as one transaction
        total_transactions = total_transactions or entry_count

    period_start = min(dates).isoformat() if dates else None
    period_end = max(dates).isoformat() if dates else None
    average_transaction_value = (total_revenue / total_transactions) if total_transactions else 0.0

    return {
        'total_revenue': total_revenue,
        'total_transactions': total_transactions,
        'entry_count': entry_count,
        'period_start': period_start,
        'period_end': period_end,
        'average_transaction_value': average_transaction_value,
    }


def generate_transaction_detail_report(
    ledger_entries: List[Dict[str, Any]]
) -> List[Dict[str, Any]]:
    """
    Create detailed report with individual transaction information.

    Args:
        ledger_entries: List of aggregated ledger entries

    Returns:
        List of dictionaries representing report rows with all detail fields

    Logic Steps:
        1. Prepare column headers for detail report
        2. Iterate through ledger entries
        3. Format each entry for tabular display
        4. Add row identifiers and sequence numbers
        5. Return formatted report data
    """
    detail_rows: List[Dict[str, Any]] = []
    for idx, entry in enumerate(ledger_entries or [], start=1):
        row = dict(entry)  # shallow copy to avoid mutating original
        # Add sequence and a stable row identifier
        row.setdefault('sequence', idx)
        row.setdefault('row_id', f'R{idx:04d}')
        detail_rows.append(row)

    return detail_rows


def export_to_excel(
    summary_report: Dict[str, Any],
    detail_report: List[Dict[str, Any]],
    output_filepath: str
) -> bool:
    """
    Export reports to a formatted Excel workbook with multiple sheets.

    Args:
        summary_report: Summary statistics dictionary from generate_summary_report()
        detail_report: List of detail rows from generate_transaction_detail_report()
        output_filepath: Path where Excel file should be saved

    Returns:
        True if export successful, False otherwise

    Logic Steps:
        1. Create new Excel workbook
        2. Add "Summary" sheet with summary_report statistics
        3. Add "Details" sheet with detail_report data
        4. Format cells (headers, numeric formatting, column widths)
        5. Apply conditional formatting if needed
        6. Save workbook to output_filepath
        7. Return success status
    """
    if Workbook is None:
        raise RuntimeError('openpyxl is required to export Excel reports')

    try:
        out_path = Path(output_filepath)
        out_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()

        # Summary sheet
        summary_ws = wb.active
        summary_ws.title = 'Summary'
        row = 1
        for key in ('total_revenue', 'total_transactions', 'entry_count', 'period_start', 'period_end', 'average_transaction_value'):
            if key in summary_report:
                summary_ws.cell(row=row, column=1, value=key)
                value = summary_report.get(key)
                # For revenue fields, write as number where possible
                if isinstance(value, (int, float)) and 'revenue' in key:
                    cell = summary_ws.cell(row=row, column=2, value=float(value))
                    cell.number_format = '$#,##0.00'
                else:
                    summary_ws.cell(row=row, column=2, value=value)
                row += 1

        _apply_summary_sheet_formatting(summary_ws)

        # Details sheet
        details_ws = wb.create_sheet('Details')
        if not detail_report:
            details_ws.cell(row=1, column=1, value='No data')
        else:
            # Headers
            headers = list(detail_report[0].keys())
            for col_idx, header in enumerate(headers, start=1):
                cell = details_ws.cell(row=1, column=col_idx, value=header)
                if Font is not None:
                    cell.font = Font(bold=True)

            # Rows
            for r_idx, row_data in enumerate(detail_report, start=2):
                for c_idx, header in enumerate(headers, start=1):
                    val = row_data.get(header)
                    cell = details_ws.cell(row=r_idx, column=c_idx, value=val)
                    # Apply currency formatting heuristically
                    if isinstance(val, (int, float)) and any(k in header.lower() for k in ('revenue', 'amount', 'total')):
                        cell.number_format = '$#,##0.00'

        _apply_detail_sheet_formatting(details_ws)

        wb.save(str(out_path))
        return True
    except Exception:
        return False


def format_currency_column(value: float) -> str:
    """
    Format monetary value as currency string.

    Args:
        value: Numeric amount

    Returns:
        Formatted currency string (e.g., "$1,234.56")
    """
    try:
        return '${:,.2f}'.format(float(value))
    except Exception:
        return str(value)


def _apply_summary_sheet_formatting(worksheet: Any) -> None:
    """
    Internal helper to apply Excel formatting to summary sheet.

    Args:
        worksheet: Excel worksheet object to format
    """
    # Apply simple formatting: bold labels, alignments and reasonable column widths
    if Font is None or Alignment is None or get_column_letter is None:
        return

    bold = Font(bold=True)
    align = Alignment(horizontal='left', vertical='center')
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = align
    # Bold the label column
    for cell in worksheet['A']:
        cell.font = bold

    # Autosize first two columns
    for col_idx in range(1, min(3, worksheet.max_column + 1)):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in worksheet[col_letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        worksheet.column_dimensions[col_letter].width = max(10, max_len + 2)


def _apply_detail_sheet_formatting(worksheet: Any) -> None:
    """
    Internal helper to apply Excel formatting to detail sheet.

    Args:
        worksheet: Excel worksheet object to format
    """
    if Font is None or Alignment is None or get_column_letter is None:
        return

    header_font = Font(bold=True)
    align = Alignment(horizontal='left', vertical='center')

    # Header row
    for cell in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=False)):
        cell.font = header_font
        cell.alignment = align

    # Autosize columns
    for col_idx in range(1, worksheet.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in worksheet[col_letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        worksheet.column_dimensions[col_letter].width = max(10, max_len + 2)


def _create_summary_sheet(
    writer: pd.ExcelWriter,
    summary_data: Dict,
    sheet_name: str
) -> None:
    """
    Creates and formats the 'Summary' tab of the report.

    Args:
        writer (pd.ExcelWriter): The pandas ExcelWriter object.
        summary_data (Dict): Data for the summary sheet.
        sheet_name (str): The name for the worksheet.
    
    Returns:
        None
    """
    kpis: List[Tuple[str, Any]] = summary_data.get('kpis', [])
    revenue_by_source_df = summary_data.get('revenue_by_source_df', pd.DataFrame())

    if sheet_name not in writer.sheets:
        writer.book.create_sheet(sheet_name)
    ws = writer.book[sheet_name]
    writer.sheets[sheet_name] = ws

    for row_idx, (label, value) in enumerate(kpis, start=1):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=value)

    kpi_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    kpi_font = Font(name='Calibri', size=14, bold=True)
    kpi_alignment = Alignment(horizontal='left', vertical='center')

    for row_idx in range(1, len(kpis) + 1):
        for col_idx in range(1, 3):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = kpi_fill
            cell.font = kpi_font
            cell.alignment = kpi_alignment

    table_start_row = len(kpis) + 2
    if not revenue_by_source_df.empty:
        revenue_by_source_df.to_excel(
            writer,
            sheet_name=sheet_name,
            startrow=table_start_row - 1,
            index=False
        )
        ws = writer.sheets[sheet_name]
        header_row = table_start_row
        header_font = Font(name='Calibri', bold=True)
        for col_idx in range(1, revenue_by_source_df.shape[1] + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.font = header_font


def _create_chart_sheet(
    writer: pd.ExcelWriter,
    ledger_df: pd.DataFrame,
    chart_sheet_name: str,
    ledger_sheet_name: str
) -> None:
    """
    Creates a bar chart from the ledger data and adds it to a new sheet.

    Args:
        writer (pd.ExcelWriter): The pandas ExcelWriter object.
        ledger_df (pd.DataFrame): The data source for the chart.
        chart_sheet_name (str): The name for the new chart worksheet.
        ledger_sheet_name (str): The name of the sheet containing the ledger data.

    Returns:
        None
    """
    revenue_by_source = pd.DataFrame()
    daily_revenue = pd.DataFrame()

    if not ledger_df.empty:
        if 'revenue_source' in ledger_df.columns and 'total_revenue' in ledger_df.columns:
            revenue_by_source = (ledger_df
                                 .groupby('revenue_source', as_index=False)['total_revenue']
                                 .sum()
                                 .rename(columns={'revenue_source': 'Revenue Source', 'total_revenue': 'Total Revenue'}))
        if 'transaction_date' in ledger_df.columns and 'total_revenue' in ledger_df.columns:
            daily_revenue = (ledger_df
                             .groupby('transaction_date', as_index=False)['total_revenue']
                             .sum()
                             .rename(columns={'transaction_date': 'Transaction Date', 'total_revenue': 'Daily Revenue'}))

    if chart_sheet_name not in writer.sheets:
        writer.book.create_sheet(chart_sheet_name)
    ws = writer.book[chart_sheet_name]
    writer.sheets[chart_sheet_name] = ws

    row_cursor = 1
    if not revenue_by_source.empty:
        revenue_by_source.to_excel(writer, sheet_name=chart_sheet_name, startrow=row_cursor - 1, index=False)
        revenue_start_row = row_cursor
        revenue_end_row = row_cursor + len(revenue_by_source)
        revenue_data_ref = Reference(ws, min_col=2, min_row=revenue_start_row, max_row=revenue_end_row)
        revenue_cat_ref = Reference(ws, min_col=1, min_row=revenue_start_row + 1, max_row=revenue_end_row)

        bar_chart = BarChart()
        bar_chart.title = 'Revenue by Source'
        bar_chart.y_axis.title = 'Total Revenue'
        bar_chart.x_axis.title = 'Revenue Source'
        bar_chart.add_data(revenue_data_ref, titles_from_data=True)
        bar_chart.set_categories(revenue_cat_ref)

        ws.add_chart(bar_chart, 'E2')
        row_cursor = revenue_end_row + 3

    if not daily_revenue.empty:
        daily_revenue.to_excel(writer, sheet_name=chart_sheet_name, startrow=row_cursor - 1, index=False)
        daily_start_row = row_cursor
        daily_end_row = row_cursor + len(daily_revenue)
        daily_data_ref = Reference(ws, min_col=2, min_row=daily_start_row, max_row=daily_end_row)
        daily_cat_ref = Reference(ws, min_col=1, min_row=daily_start_row + 1, max_row=daily_end_row)

        line_chart = LineChart()
        line_chart.title = 'Daily Revenue Trend'
        line_chart.y_axis.title = 'Daily Revenue'
        line_chart.x_axis.title = 'Date'
        line_chart.add_data(daily_data_ref, titles_from_data=True)
        line_chart.set_categories(daily_cat_ref)

        ws.add_chart(line_chart, 'E20')


def _build_summary_payload(
    summary_data: Dict,
    ledger_df: pd.DataFrame,
    discrepancies_df: pd.DataFrame
) -> Dict:
    total_revenue = summary_data.get('total_revenue')
    if total_revenue is None and not ledger_df.empty and 'total_revenue' in ledger_df.columns:
        total_revenue = ledger_df['total_revenue'].sum()

    total_transactions = summary_data.get('total_transactions')
    if total_transactions is None:
        total_transactions = summary_data.get('transactions_processed')

    discrepancies_found = summary_data.get('discrepancies_found')
    if discrepancies_found is None:
        discrepancies_found = len(discrepancies_df) if discrepancies_df is not None else 0

    date_range_value = summary_data.get('date_range')
    if date_range_value is None:
        start_date = summary_data.get('start_date')
        end_date = summary_data.get('end_date')
        if start_date and end_date:
            date_range_value = f"{start_date} to {end_date}"
        elif not ledger_df.empty and 'transaction_date' in ledger_df.columns:
            min_date = pd.to_datetime(ledger_df['transaction_date']).min()
            max_date = pd.to_datetime(ledger_df['transaction_date']).max()
            if pd.notna(min_date) and pd.notna(max_date):
                date_range_value = f"{min_date.date()} to {max_date.date()}"

    revenue_by_source_df = summary_data.get('revenue_by_source')
    if revenue_by_source_df is None:
        if not ledger_df.empty and 'revenue_source' in ledger_df.columns and 'total_revenue' in ledger_df.columns:
            revenue_by_source_df = (ledger_df
                                    .groupby('revenue_source', as_index=False)['total_revenue']
                                    .sum()
                                    .rename(columns={'revenue_source': 'Revenue Source', 'total_revenue': 'Total Revenue'}))
        else:
            revenue_by_source_df = pd.DataFrame(columns=['Revenue Source', 'Total Revenue'])
    elif isinstance(revenue_by_source_df, dict):
        revenue_by_source_df = pd.DataFrame(
            [{'Revenue Source': key, 'Total Revenue': value} for key, value in revenue_by_source_df.items()]
        )

    kpis = [
        ('Total Revenue', total_revenue if total_revenue is not None else 0),
        ('Total Transactions', total_transactions if total_transactions is not None else 0),
        ('Discrepancies Found', discrepancies_found),
        ('Date Range', date_range_value if date_range_value is not None else 'N/A')
    ]

    return {
        'kpis': kpis,
        'revenue_by_source_df': revenue_by_source_df
    }


def _format_ledger_sheet(ws, ledger_df: pd.DataFrame) -> None:
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(name='Calibri', color='FFFFFF', bold=True)
    alt_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font

    total_row_idx = ws.max_row + 1
    if ledger_df is not None and not ledger_df.empty:
        numeric_columns = [
            idx for idx, col in enumerate(ledger_df.columns, start=1)
            if pd.api.types.is_numeric_dtype(ledger_df[col])
        ]
        ws.cell(row=total_row_idx, column=1, value='TOTAL')
        for col_idx in numeric_columns:
            column_name = ledger_df.columns[col_idx - 1]
            total_value = ledger_df[column_name].sum()
            ws.cell(row=total_row_idx, column=col_idx, value=total_value)

        total_font = Font(name='Calibri', bold=True)
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=total_row_idx, column=col_idx)
            cell.font = total_font

        for row_idx in range(2, total_row_idx):
            if (row_idx - 2) % 2 == 1:
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = alt_fill


def _prepare_discrepancies(discrepancies_df: pd.DataFrame) -> pd.DataFrame:
    if discrepancies_df is None:
        return pd.DataFrame()

    discrepancies_output = discrepancies_df.copy()
    if 'reason_code' not in discrepancies_output.columns:
        if 'discrepancy_type' in discrepancies_output.columns:
            discrepancies_output['reason_code'] = discrepancies_output['discrepancy_type']
        elif 'discrepancy_reason' in discrepancies_output.columns:
            discrepancies_output['reason_code'] = discrepancies_output['discrepancy_reason']
        else:
            discrepancies_output['reason_code'] = ''

    discrepancies_output = discrepancies_output.sort_values('reason_code').reset_index(drop=True)
    return discrepancies_output


def _format_discrepancies_sheet(ws, discrepancies_df: pd.DataFrame) -> None:
    reason_col_idx = None
    for col_idx in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col_idx).value == 'reason_code':
            reason_col_idx = col_idx
            break

    header_font = Font(name='Calibri', bold=True)
    for col_idx in range(1, ws.max_column + 1):
        ws.cell(row=1, column=col_idx).font = header_font

    data_fill = PatternFill(start_color='FFE0E0', end_color='FFE0E0', fill_type='solid')
    reason_font = Font(name='Calibri', color='FF0000')

    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).fill = data_fill
        if reason_col_idx is not None:
            ws.cell(row=row_idx, column=reason_col_idx).font = reason_font


def _autosize_columns(ws) -> None:
    for col_idx in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        for cell in ws[column_letter]:
            if cell.value is None:
                continue
            cell_length = len(str(cell.value))
            if cell_length > max_length:
                max_length = cell_length
        ws.column_dimensions[column_letter].width = max(10, min(40, max_length + 2))


def _apply_font(ws, font_name: str) -> None:
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            cell.font = cell.font.copy(name=font_name)
